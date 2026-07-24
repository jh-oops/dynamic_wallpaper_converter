#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
动态壁纸 ZIP 生成器 · 本地 Web 服务
===================================
把 `transcode_for_zip.py` 的全部流水线（转码 H.264 / 改分辨率 / 抽帧 / 缩略图 /
manifest / 打包，可选动态预览图）通过一个网页前端暴露出来，任何人用浏览器即可操作，
无需命令行。

启动：
    python app.py
    # 自动打开 http://localhost:8000/ ；若未自动打开，手动访问该地址。

依赖（已在隔离 venv 安装）：flask、imageio-ffmpeg、Pillow
若 flask 缺失，脚本会尝试自动 pip 安装。
"""
import base64
import json
import os
import shutil
import subprocess
import sys
import tempfile
import threading
import webbrowser

# ---- 自举：缺 flask 时尝试安装 ----
try:
    from flask import Flask, request, jsonify
except ImportError:
    import subprocess
    print("未检测到 flask，正在自动安装…")
    subprocess.run([sys.executable, "-m", "pip", "install", "flask", "-q"], check=True)
    from flask import Flask, request, jsonify

import io

import openpyxl
import transcode_for_zip as tz
import tag_matcher as tm
import image_analyzer as ia
import vision_ai as va

HERE = os.path.dirname(os.path.abspath(__file__))
HTML_PATH = os.path.join(HERE, "web_ui.html")
LIBRARY_PATH = os.path.join(HERE, "tags_library.json")
PORT = 8000

app = Flask(__name__)


@app.route("/")
def index():
    with open(HTML_PATH, "r", encoding="utf-8") as f:
        return f.read()


@app.route("/api/package", methods=["POST"])
def package():
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify(error="请上传 mp4 文件"), 400

    name = (request.form.get("name") or os.path.splitext(f.filename)[0]).strip()
    author = (request.form.get("author") or "").strip()
    desc = (request.form.get("description") or "").strip()
    id_fixed = request.form.get("id_fixed") == "1"
    dynamic = request.form.get("dynamic") == "1"

    # 必填校验（与规范 / 网页工具一致）
    errs = []
    if not name:
        errs.append("名称 name 不能为空")
    if not author:
        errs.append("作者 author 不能为空（须与设计师账号一致，区分大小写）")
    if not desc:
        errs.append("描述 description 不能为空")
    if errs:
        return jsonify(error="；".join(errs)), 400

    # 落盘上传文件到临时目录
    tmp = tempfile.mkdtemp(prefix="wp_up_")
    src = os.path.join(tmp, os.path.basename(f.filename))
    f.save(src)
    try:
        res = tz.process_one(src, name, author, desc, id_fixed, False, dynamic)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

    if not res["ok"]:
        return jsonify(error="处理失败：" + res.get("error", "未知错误")), 500

    return jsonify({
        "zip_name": res["zip_name"],
        "zip_b64": base64.b64encode(res["zip_bytes"]).decode("ascii"),
        "dyn_name": res["dyn_name"],
        "dyn_b64": base64.b64encode(res["dyn_bytes"]).decode("ascii") if res["dyn_bytes"] else None,
        "report": res["report"],
    })


@app.route("/api/tags/library", methods=["GET"])
def get_library():
    """返回当前标签库（分类 + 标签）。"""
    return jsonify(tm.load_library(LIBRARY_PATH))


@app.route("/api/tags/library", methods=["PUT"])
def put_library():
    """保存标签库。简单校验必须包含 categories 与 tags 数组。"""
    data = request.get_json(force=True, silent=True) or {}
    if not isinstance(data.get("categories"), list) or not isinstance(data.get("tags"), list):
        return jsonify(error="标签库必须包含 categories 和 tags 数组"), 400
    try:
        with open(LIBRARY_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return jsonify(ok=True)
    except Exception as e:
        return jsonify(error=f"保存失败：{e}"), 500


@app.route("/api/analyze", methods=["POST"])
def analyze():
    """上传图片或视频，返回建议的 category + tags。

    识别逻辑：基础为文件名规则匹配；若文件是图片，直接读像素分析；
    若是视频，抽首帧后读像素分析。图像特征与文件名打分合并，
    每个标签标注来源（filename / image / both），并附带视觉特征供前端展示“识别依据”。
    """
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify(error="请上传文件"), 400
    lib = tm.load_library(LIBRARY_PATH)
    ext = os.path.splitext(f.filename)[1].lower()

    image_scores = None
    image_cat_scores = None
    features = None
    analyzed_kind = None
    recognition_source = None
    tmp = None
    try:
        if ext in IMAGE_EXTS:
            data = f.read()
            analyzed_kind = "image"
            image_scores, image_cat_scores, features, recognition_source = va.analyze(data, f.filename, library=lib)
        elif ext == ".mp4":
            tmp = tempfile.mkdtemp(prefix="wp_an_")
            src = os.path.join(tmp, os.path.basename(f.filename))
            f.save(src)
            frame = os.path.join(tmp, "frame.jpg")
            _extract_first_frame(src, frame)
            with open(frame, "rb") as fh:
                data = fh.read()
            analyzed_kind = "video-frame"
            image_scores, image_cat_scores, features, recognition_source = va.analyze(data, f.filename, library=lib)
        else:
            # 其他类型：尝试按图片解析，失败则仅用文件名
            try:
                data = f.read()
                image_scores, image_cat_scores, features, recognition_source = va.analyze(data, f.filename, library=lib)
                analyzed_kind = "image"
            except Exception:
                features = None
    except Exception as e:
        features = None
        image_scores = None
        image_cat_scores = None
        recognition_source = None
    finally:
        if tmp:
            shutil.rmtree(tmp, ignore_errors=True)

    res = tm.suggest(f.filename, lib, top_n=5,
                     image_scores=image_scores, image_cat_scores=image_cat_scores)
    res["analyzed_kind"] = analyzed_kind
    res["features"] = features
    res["recognition_source"] = recognition_source
    return jsonify(res)


IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".gif", ".tiff"}


def _extract_first_frame(src, dst_jpg):
    """用 ffmpeg 抽视频首帧为 JPG（供图像理解使用）。"""
    exe = tz.ffmpeg_exe()
    subprocess.run(
        [exe, "-y", "-i", src, "-vf", "select=eq(n\\,0)", "-frames:v", "1", "-q:v", "3", dst_jpg],
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, check=True,
    )
    return dst_jpg


@app.route("/api/crop", methods=["POST"])
def crop():
    """上传静态图片，按指定目标尺寸和模式裁剪后返回 JPEG（base64）。

    仅支持静态图片（jpg/png/webp/bmp/gif/tiff）。
    mode: cover（缩放铺满居中裁剪）/ crop（中心裁切）。
    可选 max_size_kb 控制输出大小（0=不限制）。
    """
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify(error="请上传图片"), 400

    try:
        target_w = int(request.form.get("target_w", 1080))
        target_h = int(request.form.get("target_h", 2400))
    except ValueError:
        return jsonify(error="target_w / target_h 必须是整数"), 400
    if target_w < 1 or target_h < 1:
        return jsonify(error="目标尺寸必须大于 0"), 400

    mode = request.form.get("mode", "cover")
    if mode not in ("cover", "crop"):
        return jsonify(error='mode 只能是 "cover" 或 "crop"'), 400
    try:
        max_size_kb = int(request.form.get("max_size_kb", 0) or 0)
    except ValueError:
        return jsonify(error="max_size_kb 必须是整数"), 400
    if max_size_kb < 0:
        max_size_kb = 0

    try:
        offset_x = float(request.form.get("offset_x", "0.5"))
        offset_y = float(request.form.get("offset_y", "0.5"))
        scale = float(request.form.get("scale", "1.0"))
    except ValueError:
        return jsonify(error="offset_x / offset_y / scale 必须是数字"), 400
    offset_x = max(0.0, min(1.0, offset_x))
    offset_y = max(0.0, min(1.0, offset_y))
    scale = max(1.0, scale)

    ext = os.path.splitext(f.filename)[1].lower()
    is_image = ext in IMAGE_EXTS
    if not is_image:
        return jsonify(error="仅支持静态图片，请上传 JPG / PNG / WebP / GIF / BMP 等"), 400
    base = os.path.splitext(os.path.basename(f.filename))[0]

    tmp = tempfile.mkdtemp(prefix="wp_crop_")
    src = os.path.join(tmp, os.path.basename(f.filename))
    f.save(src)
    try:
        dst = os.path.join(tmp, "cropped.jpg")
        tz.crop_image(src, dst, target_w, target_h, mode,
                      offset_x=offset_x, offset_y=offset_y, scale=scale)
        # 大小限制：若超过 max_size_kb，逐步降低 JPEG 质量重试
        within = True
        if max_size_kb > 0:
            limit = max_size_kb * 1024
            q = 88
            while os.path.getsize(dst) > limit and q >= 15:
                tz.crop_image(src, dst, target_w, target_h, mode, quality=q,
                              offset_x=offset_x, offset_y=offset_y, scale=scale)
                q -= 8
            within = os.path.getsize(dst) <= limit
        with open(dst, "rb") as fh:
            data_bytes = fh.read()
        return jsonify({
            "ok": True,
            "kind": "image",
            "file_name": f"{base}_{target_w}x{target_h}_{mode}.jpg",
            "file_b64": base64.b64encode(data_bytes).decode("ascii"),
            "mime": "image/jpeg",
            "target_w": target_w,
            "target_h": target_h,
            "mode": mode,
            "size_kb": round(len(data_bytes) / 1024, 1),
            "max_size_kb": max_size_kb,
            "within_limit": within,
        })
    except Exception as e:
        return jsonify(error=f"裁剪失败：{e}"), 500
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


# ── 视觉识别后端配置 ──
@app.route("/api/vision_config", methods=["GET"])
def get_vision_config():
    cfg = va.load_config()
    # 不回传 api_key 明文，仅告知是否已配置
    masked = dict(cfg)
    if masked.get("api_key"):
        masked["api_key"] = "******" if len(masked["api_key"]) > 6 else ""
        masked["api_key_set"] = True
    else:
        masked["api_key_set"] = False
    return jsonify(masked)


@app.route("/api/vision_config", methods=["PUT"])
def put_vision_config():
    data = request.get_json(force=True, silent=True) or {}
    cfg = va.load_config()
    for k in ("backend", "base_url", "api_key", "model", "prompt"):
        if k in data:
            cfg[k] = data[k]
    cfg = va.save_config(cfg)
    return jsonify(ok=True, backend=cfg.get("backend"))


# ── 批量静态壁纸打包 ──
# Excel 列名（中/英别名）→ 内部键
COLUMN_ALIASES = {
    "designer_id": ["设计师id", "设计师ID", "设计师_id", "designer_id", "designerid", "作者id", "author_id"],
    "name":        ["英文标题", "名称", "name", "壁纸名称", "文件名", "filename", "title"],
    "desc":        ["英文描述", "简介", "描述", "desc", "description", "introduce"],
    "price":       ["资源价格", "价格", "定价", "price", "cost"],
    "category":    ["二级分类", "分类", "category", "类型", "type"],
    "tags":        ["标签", "tags", "tag"],
    "quality":     ["质量", "质量标准", "quality", "grade", "等级"],
}


def _resolve_columns(header):
    norm = {}
    for i, h in enumerate(header):
        if h is None:
            continue
        norm[str(h).strip().lower()] = i + 1  # 1-based 列号
    colmap = {}
    for key, aliases in COLUMN_ALIASES.items():
        for a in aliases:
            if a.lower() in norm:
                colmap[key] = norm[a.lower()]
                break
    return colmap


def _ensure_column(ws, colmap, key, header_label):
    """若列不存在则追加，返回列号（1-based）。"""
    if key in colmap:
        return colmap[key]
    max_col = ws.max_column
    new_col = max_col + 1
    ws.cell(1, new_col, header_label)
    colmap[key] = new_col
    return new_col


def _pick_top_tags(tag_scores, lib, top_n=5):
    if not tag_scores:
        return []
    known = {t["name"] for t in lib.get("tags", [])}
    ranked = sorted(
        ((n, s) for n, s in tag_scores.items() if n in known),
        key=lambda kv: -kv[1],
    )
    return [n for n, _ in ranked[:top_n]]


def _estimate_quality(features):
    """基于像素特征给出质量等级 A/B/C/D。"""
    if not features:
        return "B"
    score = 50
    w, h = features.get("size") or (0, 0)
    # 分辨率
    if min(w, h) >= 2160:
        score += 20
    elif min(w, h) >= 1080:
        score += 10
    else:
        score -= 15
    # 细节密度
    edge = features.get("edge_density", 0)
    if edge > 35:
        score += 15
    elif edge > 20:
        score += 8
    else:
        score -= 5
    # 对比度
    contrast = features.get("contrast", 0)
    if contrast > 45:
        score += 10
    elif contrast < 20:
        score -= 10
    # 色彩丰富度
    colorfulness = features.get("colorfulness", 0)
    if colorfulness > 50:
        score += 10
    elif colorfulness > 30:
        score += 5
    # 饱和度
    if features.get("mean_sat", 0) > 0.4:
        score += 5
    if score >= 85:
        return "A"
    if score >= 70:
        return "B"
    if score >= 55:
        return "C"
    return "D"


@app.route("/api/batch_package", methods=["POST"])
def batch_package():
    """批量静态壁纸打包：
    上传 Excel（含 名称/简介/定价/分类，理论由用户先提供）+ 所有图片（文件名=名称）+ 设计师ID
    → 对每张图跑识别引擎（offline 或 ai）→ 把识别标签写入 标签 列、设计师ID 写入所有行
    → 输出 zip（更新后的 Excel + 所有图片）。
    可选表单字段：designer_id、backend（覆盖配置，offline/ai）。
    """
    excel = request.files.get("excel")
    if not excel or not excel.filename:
        return jsonify(error="请上传 Excel 文件（.xlsx）"), 400
    designer_id = (request.form.get("designer_id") or "").strip()
    backend = (request.form.get("backend") or "").strip() or None
    normalize = request.form.get("normalize") == "1"
    images = request.files.getlist("images")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(excel.read()))
    except Exception as e:
        return jsonify(error=f"Excel 读取失败：{e}"), 400
    ws = wb.active

    header = [c.value for c in ws[1]]
    colmap = _resolve_columns(header)
    if "name" not in colmap:
        return jsonify(error="Excel 缺少“英文标题”列（用于与图片文件名对应）"), 400

    # 确保 设计师id / 标签 / 质量 列存在
    _ensure_column(ws, colmap, "designer_id", "设计师id")
    _ensure_column(ws, colmap, "tags", "标签")
    _ensure_column(ws, colmap, "quality", "质量")

    lib = tm.load_library(LIBRARY_PATH)
    cfg = va.load_config()
    if backend:
        cfg = dict(cfg)
        cfg["backend"] = backend

    # 图片按“名称（去扩展名）”索引
    img_by_base = {}
    for im in images:
        base = os.path.splitext(os.path.basename(im.filename))[0]
        img_by_base[base] = im

    out = tempfile.mkdtemp(prefix="wp_batch_")
    img_dir = os.path.join(out, "images")
    os.makedirs(img_dir, exist_ok=True)

    preview = []
    matched = 0
    unmatched = 0
    lib_cat_names = {c["name"] for c in lib.get("categories", [])}

    for r in range(2, ws.max_row + 1):
        name_val = ws.cell(r, colmap["name"]).value
        if name_val is None or str(name_val).strip() == "":
            continue
        name = str(name_val).strip()

        # 设计师 ID：填充到所有行的 设计师id 列
        if designer_id:
            ws.cell(r, colmap["designer_id"]).value = designer_id

        base = os.path.splitext(name)[0]
        im = img_by_base.get(base)
        if not im:
            # 也尝试忽略大小写/空白
            im = img_by_base.get(base.strip())
        if not im:
            unmatched += 1
            preview.append({
                "name": name,
                "category": _cell(ws, r, colmap.get("category")),
                "tags": "",
                "image_matched": False,
                "source": None,
            })
            continue

        # 保存原图到 images/（保留原名，确保与 名称 对应）
        # 注意：必须先 read() 再写盘——FileStorage.save() 会消耗流，之后 read() 为空
        data = im.read()
        # 若启用标准化，统一裁剪为 1080x2160 cover（静态壁纸标准尺寸）
        if normalize:
            tmp_src = os.path.join(out, f"_norm_{r}.jpg")
            tmp_dst = os.path.join(out, f"_norm_{r}_out.jpg")
            with open(tmp_src, "wb") as fh:
                fh.write(data)
            try:
                tz.crop_image(tmp_src, tmp_dst, 1080, 2160, "cover")
                with open(tmp_dst, "rb") as fh:
                    data = fh.read()
            except Exception:
                pass
            finally:
                for p in (tmp_src, tmp_dst):
                    if os.path.exists(p):
                        os.remove(p)
        img_path = os.path.join(img_dir, os.path.basename(im.filename))
        with open(img_path, "wb") as fh:
            fh.write(data)

        try:
            ts, cs, feats, src = va.analyze(data, im.filename, cfg=cfg, library=lib)
        except Exception as e:
            ts, cs, feats, src = {}, {}, None, f"error:{e}"

        tags = _pick_top_tags(ts, lib, top_n=5)
        ws.cell(r, colmap["tags"]).value = "、".join(tags)

        # 若 质量 列为空，按图像特征自动识别填充
        if "quality" in colmap:
            existing_q = ws.cell(r, colmap["quality"]).value
            if existing_q is None or str(existing_q).strip() == "":
                ws.cell(r, colmap["quality"]).value = _estimate_quality(feats)

        # 若 分类 列为空且识别给出了有效分类，则补全
        if "category" in colmap:
            existing = ws.cell(r, colmap["category"]).value
            if (existing is None or str(existing).strip() == "") and cs:
                top_cat = sorted(cs.items(), key=lambda kv: -kv[1])[0][0]
                if top_cat in lib_cat_names:
                    ws.cell(r, colmap["category"]).value = top_cat

        matched += 1
        preview.append({
            "name": name,
            "category": _cell(ws, r, colmap.get("category")),
            "tags": "、".join(tags),
            "image_matched": True,
            "source": src,
        })

    # 写出 Excel
    xlsx_path = os.path.join(out, "wallpapers.xlsx")
    wb.save(xlsx_path)

    # 打 zip
    zip_path = os.path.join(out, "batch_wallpapers.zip")
    import zipfile
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(xlsx_path, "wallpapers.xlsx")
        for fn in os.listdir(img_dir):
            zf.write(os.path.join(img_dir, fn), os.path.join("images", fn))

    with open(zip_path, "rb") as fh:
        zip_bytes = fh.read()

    # 生成前端可预览的 Excel 行数据（按表头）
    excel_preview = []
    headers = [c.value for c in ws[1]]
    for r in range(2, ws.max_row + 1):
        name_val = ws.cell(r, colmap["name"]).value
        if name_val is None or str(name_val).strip() == "":
            continue
        row_dict = {}
        for c_idx, h in enumerate(headers, start=1):
            row_dict[str(h) if h is not None else f"列{c_idx}"] = ws.cell(r, c_idx).value or ""
        excel_preview.append(row_dict)

    shutil.rmtree(out, ignore_errors=True)
    return jsonify({
        "ok": True,
        "zip_name": "batch_wallpapers.zip",
        "zip_b64": base64.b64encode(zip_bytes).decode("ascii"),
        "stats": {
            "total_rows": len(preview),
            "matched": matched,
            "unmatched": unmatched,
            "backend": cfg.get("backend", "offline"),
        },
        "preview": preview,
        "excel_preview": excel_preview,
    })


def _cell(ws, row, col):
    if not col:
        return ""
    v = ws.cell(row, col).value
    return str(v).strip() if v is not None else ""


if __name__ == "__main__":
    url = f"http://localhost:{PORT}/"
    # 1 秒后自动打开浏览器
    threading.Timer(1.0, lambda: webbrowser.open(url)).start()
    print(f"动态壁纸生成器已启动：{url}")
    print("（关闭此终端窗口即停止服务）")
    app.run(host="127.0.0.1", port=PORT, debug=False)
